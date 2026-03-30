from openpyxl import load_workbook
from datetime import datetime

kinmu_file = "[末松亜域] 勤務表.xlsx"
kotsuhi_file = "[末松亜域] 交通費明細書.xlsx"
output_file = "交通費明細書_完成.xlsx"

kinmu_sheet_name = "3月勤務表"
kotsuhi_sheet_name = "(2026年3月)"
VISIT_PLACE = "品川オフィス"
PURPOSE = "出社"
ROUTE = "JR上野東京ライン"
FROM_STATION = "赤羽"
TO_STATION = "品川"
TRIP_TYPE = "往復"
FARE = "9620円"
COST_TYPE = "定期"

weekday_map = {
0:"月",1:"火",2:"水",3:"木",4:"金",5:"土",6:"日"
}

wb_kinmu = load_workbook(kinmu_file, data_only=True)
ws_kinmu = wb_kinmu[kinmu_sheet_name]

wb_kotsuhi = load_workbook(kotsuhi_file)
print(wb_kotsuhi.sheetnames)
ws_kotsuhi = wb_kotsuhi[kotsuhi_sheet_name]

office_days=[]

for row in range(5,33):

    work_date = ws_kinmu[f"B{row}"].value
    location = ws_kinmu[f"I{row}"].value

    if location == "出社" and work_date:
        office_days.append(work_date)

start_row=12

for i,work_date in enumerate(office_days):

    target_row=start_row+i

    if isinstance(work_date,datetime):
        dt=work_date
    else:
        dt=datetime.strptime(str(work_date),"%Y/%m/%d")

    ws_kotsuhi[f"A{target_row}"]=f"{dt.month}/{dt.day}"
    ws_kotsuhi[f"B{target_row}"]=weekday_map[dt.weekday()]
    ws_kotsuhi[f"C{target_row}"]=VISIT_PLACE
    ws_kotsuhi[f"F{target_row}"]=PURPOSE
    ws_kotsuhi[f"H{target_row}"]=ROUTE
    ws_kotsuhi[f"I{target_row}"]=FROM_STATION
    ws_kotsuhi[f"J{target_row}"]=TO_STATION
    ws_kotsuhi[f"K{target_row}"]=TRIP_TYPE

    if i==0:
        ws_kotsuhi[f"L{target_row}"]=FARE
        ws_kotsuhi[f"N{target_row}"]=COST_TYPE

wb_kotsuhi.save(output_file)

print("交通費明細書作成完了")
